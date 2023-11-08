"""App to run the main script."""
# Third-party imports
import openpyxl

# File path for the input excel document.
# 'S:\Customers\23 V1Fiber\23.065.001 V1 Corning Lima\Scranton\Weekly Lima Report Data\Bain Data Comparison.xlsx'
# Initial file paths
file_input = r'S:\Customers\23 V1Fiber\23.065.001 V1 Corning Lima\Scranton\Weekly Lima Report Data\Bain Data Comparison.xlsx'
file_output = r''

# Initial lists for data
list_our_data = []
list_corning_data = []

# Open the workbook and create sheet objects.
wb = openpyxl.load_workbook(file_input)
ws_our_data = wb['Sheet1']
ws_corning_data = wb['Sheet2']

# Set the start and end row to
# iterate through ws_our_data
start_row = 2
end_row = ws_our_data.max_row

# Initial variables to build the address
mail_add = ''
unit_number = ''
municipality = ''
zip_code = ''
lat = ''
long = ''
address_full = ''

# Iterate through every row in the sheet
# containing our data. Format the address.
# Store in an array.
for row in range(start_row, end_row+1):
    mail_add = ws_our_data.cell(row=row,column=2).value
    unit_number = ws_our_data.cell(row=row,column=3).value
    municipality = ws_our_data.cell(row=row,column=4).value
    zip_code = ws_our_data.cell(row=row,column=5).value
    lat = ws_our_data.cell(row=row,column=7).value
    long = ws_our_data.cell(row=row,column=8).value
    address_full = f'{mail_add} unit {unit_number}, {municipality} {zip_code}; {lat} {long}'
    #address_full = f'{mail_add} unit {unit_number}, {municipality} {zip_code}'
    list_our_data.append(address_full.lower())

# Repeat above, but for corning data.
# Set the start and end row to
# iterate through ws_our_data
start_row = 2
end_row = ws_corning_data.max_row

# Iterate through every row in the sheet
# containing our data. Format the address.
# Store in an array.
for row in range(start_row, end_row+1):
    mail_add = ws_corning_data.cell(row=row,column=3).value
    unit_number = ws_corning_data.cell(row=row,column=4).value
    municipality = ws_corning_data.cell(row=row,column=5).value
    zip_code = ws_corning_data.cell(row=row,column=6).value
    lat = ws_corning_data.cell(row=row,column=7).value
    long = ws_corning_data.cell(row=row,column=8).value
    address_full = f'{mail_add} unit {unit_number}, {municipality} {zip_code}; {lat} {long}'
    #address_full = f'{mail_add} unit {unit_number}, {municipality} {zip_code}'
    list_corning_data.append(address_full.lower())

# Test to check the length our list. Should be 8,262.
# Also provide sample output.
print('--------------------------')
list_length = len(list_our_data)
print(f'The list length is {list_length}')
print('Sample output:')
print('--------------------------')
for i in range(0,5):
    print(list_our_data[i])

# Test to check the length our list. Should be 36,155.
# Also provide sample output.
print('--------------------------')
list_length = len(list_corning_data)
print(f'The list length is {list_length}')
print('Sample output:')
print('--------------------------')
for i in range(0,5):
    print(list_corning_data[i])
print('--------------------------')
# Checks on the initial list output are complete.

# Store a sorted copy of each list.
list_sort_our_data = sorted(list_our_data)
list_sort_corning_data = sorted(list_corning_data)
"""
# Test to see if the script can find mathing
# addresses
i = 0
for address in list_sort_our_data:
    if address in list_sort_corning_data:
        print(f'Match found for: {address}')
        i += 1
        if i > 10:
            break
print('--------------------------')
"""
# Initial variables for matched 
# and unmatched lists.
list_unmatched_addresses = []
list_matched_addresses_our_data = []

# Initial variables to store lists with no lat and long.
list_sort_split_our_data = []
list_sort_split_corning_data = []

# Populate list_sort_split_our_data
for address in list_sort_our_data:
    list_sort_split_our_data.append(address.split(';')[0])
    #print(address.split(';')[0])

# Populated the list_sort_split_corning_data
for address in list_sort_corning_data:
    list_sort_split_corning_data.append(address.split(';')[0])

# Iterate through the sorted lists.
# to look for matching addresses.
# Store both matched and unmatched addresses.
# If an address is matched, also remove it from 
# the sorted list of corning data.
for index, address in enumerate(list_sort_split_our_data):
    if address in list_sort_split_corning_data:
        index_corning_data = list_sort_split_corning_data.index(address)
        # Append matched address to list_matched_address.
        list_matched_addresses_our_data.append(list_sort_our_data[index])
        # Remove matched address from corning data.
        list_sort_corning_data.remove(list_sort_corning_data[index_corning_data])
        list_sort_split_corning_data.remove(address)
    else:
        list_unmatched_addresses.append(list_sort_our_data[index])

print(f'Length of matched addresses: {len(list_matched_addresses_our_data)}')
print(f'Length of un-matched addresses: {len(list_unmatched_addresses)}')
print(f'Length of corning addresses: {len(list_sort_corning_data)}')
print('--------------------------')

# Go through the unmatched address list,
# matched address list, and sorted
# corning data and write it to a new excel file.
wb_output = openpyxl.Workbook()



